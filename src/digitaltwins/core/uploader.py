import os
import json
import logging
from pathlib import Path
from typing import Optional

import psycopg2
import openpyxl

from dotenv import load_dotenv
load_dotenv()

from ..utils.config_loader import is_truthy

from ..postgres.uploader import Uploader as PostgresUploader
from ..irods.uploader import Uploader as IRODSUploader
from ..minio.uploader import Uploader as MinioUploader

logger = logging.getLogger(__name__)

# Mapping from .xlsx base filenames (without extension) to Postgres table names.
# Filenames not in this map are skipped during DB insertion.
XLSX_TABLE_MAP = {
    "dataset_description": "dataset_description",
    "subjects": "subject",
    "samples": "sample",
    "manifest": "manifest",
}

# Tables whose multiple Excel rows should be merged into a single DB record.
# Fields with multiple values across rows are combined into lists (Postgres arrays).
_MERGE_TABLES = {"dataset_description"}

# Columns in dataset_description that are defined as Postgres arrays (e.g., character varying[]).
_ARRAY_COLUMNS = {
    "keywords", "funding", "study_organ_system", "study_approach",
    "study_technique", "contributor_name", "contributor_orcid",
    "contributor_affiliation", "contributor_role", "identifier_description",
    "relation_type", "identifier", "identifier_type"
}

def _normalize_key(key: str) -> str:
    """Normalize an Excel header or metadata element string."""
    if not key:
        return ""
    return str(key).strip().lower().replace(" ", "_").replace("-", "_")

def _parse_excel_to_dicts(file_path: str, mapping: Optional[dict] = None) -> list[dict]:
    """Parse an .xlsx file and return a list of row-dicts.

    The first row is treated as the header.  Empty rows are skipped.
    Only cells with non-None values are included in the dict.
    If a mapping is provided, headers are translated using the map.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    raw_headers = [str(h).strip() if h else None for h in rows[0]]
    mapped_headers = []
    for h in raw_headers:
        if not h:
            mapped_headers.append(None)
            continue
        
        norm_h = _normalize_key(h)
        if mapping and norm_h in mapping:
            mapped_headers.append(mapping[norm_h])
        else:
            mapped_headers.append(norm_h)

    result = []
    for row in rows[1:]:
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue
        record = {}
        for header, cell in zip(mapped_headers, row):
            if header and cell is not None:
                record[header] = cell
        if record:
            result.append(record)
    return result

def _parse_transposed_excel(file_path: str, mapping: Optional[dict] = None) -> list[dict]:
    """Parse a transposed .xlsx file like dataset_description.xlsx.
    
    Returns a list with a single dict representing the parsed record.
    Keys are derived from the first column, and values are combined
    from columns starting at index 3 (Value, Value 1, etc.).
    If there are multiple values, they are combined into a Python list
    (which psycopg2 automatically converts to a Postgres array format).
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    record = {}
    for row in rows[1:]:
        if not row: continue
        
        # 'Metadata element' is in col 0
        raw_key = row[0]
        if raw_key is None or str(raw_key).strip() == "":
            continue
            
        norm_key = _normalize_key(raw_key)
        if mapping and norm_key in mapping:
            key = mapping[norm_key]
        else:
            key = norm_key
            if key == "type":
                key = "dataset_type"
            
        # Extract values starting from column 3 ('Value', 'Value 2', etc.)
        # Exclude None or empty strings
        values = [str(v).strip() for v in row[3:] if v is not None and str(v).strip() != ""]
        if not values:
            continue
            
        # If the column expects a Postgres array, pass a Python list unconditionally.
        # Otherwise, pass a scalar string. If multiple values exist for a scalar column,
        # combine them into a single string.
        if key in _ARRAY_COLUMNS:
            record[key] = values
        else:
            record[key] = ", ".join(values)

    return [record] if record else []




class Uploader(object):
    def __init__(self):
        self._postgres_enabled = is_truthy(os.getenv("POSTGRES_ENABLED"))
        self._gen3_enabled = is_truthy(os.getenv("GEN3_ENABLED"))
        self._irods_enabled = is_truthy(os.getenv("IRODS_ENABLED"))
        self._minio_enabled = is_truthy(os.getenv("MINIO_ENABLED"))

        if self._postgres_enabled and self._gen3_enabled:
            raise ValueError("Metadata service conflict. Only one of 'postgres' or 'gen3' can be enabled")

        self._postgres_uploader = None
        self._gen3_uploader = None
        self._irods_uploader = None

        if self._postgres_enabled:
            self._postgres_uploader = PostgresUploader()
        else:
            self._postgres_uploader = None

        if self._irods_enabled:
            self._irods_uploader = IRODSUploader()
        else:
            self._irods_uploader = None

        if self._minio_enabled:
            self._minio_uploader = MinioUploader()
        else:
            self._minio_uploader = None

        self._element_mapping = self._load_element_mapping()

    def _load_element_mapping(self) -> dict[str, dict[str, str]]:
        """Load SDS element to Postgres column mapping from Excel.
        
        Returns a dictionary: {sheet_name: {normalized_element_name: db_column_name}}
        """
        mapping_path = Path(__file__).parent.parent / "resources" / "version_2_0_0" / "element_mapping.xlsx"
        if not mapping_path.exists():
            logger.warning("Element mapping file not found at %s", mapping_path)
            return {}

        try:
            wb = openpyxl.load_workbook(str(mapping_path), read_only=True, data_only=True)
            mapping = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                
                headers = [str(h).strip().lower() if h else None for h in rows[0]]
                # We expect columns like 'element' and 'V2.0.0'
                try:
                    elem_idx = headers.index("element")
                except ValueError:
                    # Some sheets might use different casing or 'Element'
                    try:
                        elem_idx = headers.index("element")
                    except ValueError:
                        # Skip sheets without 'element' column
                        continue
                
                try:
                    db_idx = headers.index("v2.0.0")
                except ValueError:
                    # Skip sheets without 'V2.0.0' column
                    continue

                sheet_map = {}
                for row in rows[1:]:
                    elem_name = row[elem_idx]
                    db_col = row[db_idx]
                    if elem_name and db_col:
                        sheet_map[_normalize_key(str(elem_name))] = str(db_col).strip()
                mapping[sheet_name] = sheet_map
            wb.close()
            return mapping
        except Exception as e:
            logger.error("Failed to load element mapping: %s", e)
            return {}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def upload_dataset(
        self,
        dataset_path: str,
        category: str,
        save_json: bool = False,
    ) -> str:
        """Upload a SPARC SDS dataset to Postgres and MinIO.

        The function performs the following steps atomically:
        1. Parses all ``.xlsx`` metadata files in *dataset_path*.
        2. Inserts the parsed metadata into the corresponding Postgres
           tables within a single transaction.
        3. Uploads the entire *dataset_path* folder to MinIO under
           ``<category>/<dataset_uuid>/``.
        4. If the MinIO upload fails the Postgres transaction is rolled
           back so no orphaned metadata is left behind.

        Args:
            dataset_path: Path to the root of the SDS dataset directory.
            category: Logical category for the dataset (also used as the
                MinIO bucket name).
            save_json: When ``True``, each parsed ``.xlsx`` file is also
                saved as a ``.json`` file alongside the original.

        Returns:
            The generated ``dataset_uuid`` as a string.

        Raises:
            FileNotFoundError: If *dataset_path* does not exist.
            RuntimeError: If the MinIO upload fails (Postgres is rolled
                back automatically).
        """
        dataset_dir = Path(dataset_path).resolve()
        if not dataset_dir.is_dir():
            raise FileNotFoundError(f"Dataset path does not exist: {dataset_dir}")
        dataset_name = dataset_dir.name

        # ── 1. Parse all .xlsx metadata files ─────────────────────────
        xlsx_files = sorted(dataset_dir.glob("*.xlsx"))
        if not xlsx_files:
            logger.warning("No .xlsx files found in %s", dataset_dir)

        parsed_metadata: dict[str, list[dict]] = {}
        for xlsx_path in xlsx_files:
            base_name = xlsx_path.stem  # e.g. "dataset_description"
            mapping = self._element_mapping.get(base_name)

            if base_name in ["dataset_description", "code_description"]:
                records = _parse_transposed_excel(str(xlsx_path), mapping=mapping)
            else:
                records = _parse_excel_to_dicts(str(xlsx_path), mapping=mapping)
            
            parsed_metadata[base_name] = records
            logger.info("Parsed %d record(s) from %s", len(records), xlsx_path.name)

            # Optionally save as JSON
            if save_json and records:
                json_path = xlsx_path.with_suffix(".json")
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(records, f, indent=2, default=str)
                logger.info("Saved JSON to %s", json_path)

        # ── 2. Postgres upload (transactional) ────────────────────────
        conn: Optional[psycopg2.extensions.connection] = None
        try:
            if self._postgres_enabled and self._postgres_uploader:
                conn = psycopg2.connect(
                    host=self._postgres_uploader._host,
                    port=self._postgres_uploader._port,
                    database=self._postgres_uploader._database,
                    user=self._postgres_uploader._user,
                    password=self._postgres_uploader._password,
                )
                conn.autocommit = False
                cur = conn.cursor()

                # 2a. Insert into `dataset` table
                # We let PostgreSQL generate the UUID using `public.uuid_generate_v1()`
                cur.execute(
                    "INSERT INTO dataset (category, dataset_name) "
                    "VALUES (%s, %s) RETURNING dataset_uuid",
                    (category, dataset_name),
                )
                dataset_uuid_str = str(cur.fetchone()[0])
                logger.info("Inserted dataset record and received UUID: %s", dataset_uuid_str)

                # 2b. Insert metadata rows from each parsed file
                # We need to track subject/sample IDs → UUIDs for dataset_mapping
                subject_id_to_uuid: dict[str, str] = {}
                sample_mappings: list[tuple[str, str, str]] = []

                for base_name, records in parsed_metadata.items():
                    table_name = XLSX_TABLE_MAP.get(base_name)
                    if table_name is None:
                        logger.debug("Skipping %s.xlsx — no matching table", base_name)
                        continue
                    if not records:
                        continue

                    self._insert_metadata_rows(
                        cur,
                        table_name,
                        base_name,
                        records,
                        dataset_uuid_str,
                        subject_id_to_uuid,
                        sample_mappings,
                    )

                # 2c. Insert dataset_mapping rows
                if subject_id_to_uuid and sample_mappings:
                    self._insert_dataset_mapping(
                        cur,
                        dataset_uuid_str,
                        subject_id_to_uuid,
                        sample_mappings,
                    )

            # ── 3. MinIO upload ───────────────────────────────────────
            if self._minio_enabled and self._minio_uploader:
                bucket_name = category
                # Create bucket if it doesn't exist
                if not self._minio_uploader.bucket_exists(bucket_name):
                    self._minio_uploader.s3_client.create_bucket(Bucket=bucket_name)
                    logger.info("Created MinIO bucket: %s", bucket_name)

                success = self._minio_uploader.upload_folder(
                    folder_path=str(dataset_dir),
                    bucket_name=bucket_name,
                    prefix=dataset_uuid_str,
                )
                if not success:
                    raise RuntimeError(
                        f"MinIO upload failed for dataset {dataset_uuid_str}. "
                        "Postgres transaction will be rolled back."
                    )
                logger.info(
                    "Uploaded dataset to MinIO: %s/%s/", bucket_name, dataset_uuid_str
                )

            # ── 4. Commit Postgres on success ─────────────────────────
            if conn:
                conn.commit()
                logger.info("Postgres transaction committed for dataset %s", dataset_uuid_str)

        except Exception:
            # Rollback Postgres if anything went wrong
            if conn:
                conn.rollback()
                logger.error(
                    "Postgres transaction rolled back for dataset %s", dataset_uuid_str
                )
            raise
        finally:
            if conn:
                conn.close()

        return dataset_uuid_str

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _insert_metadata_rows(
        cur,
        table_name: str,
        base_name: str,
        records: list[dict],
        dataset_uuid: str,
        subject_id_to_uuid: dict,
        sample_mappings: list,
    ) -> None:
        """Insert parsed metadata rows into *table_name*.

        Handles three categories of tables:
        * **subject / sample** — inserts each row and captures the
          DB-generated UUID for later ``dataset_mapping`` population.
        * **merge tables** (``dataset_description``) — multiple Excel
          rows are merged into a single DB record; repeated fields
          become Postgres arrays.
        * **everything else** (``manifest``) — each row is inserted
          individually with dynamic column mapping.
        """
        if table_name in _MERGE_TABLES:
            # Merge all rows into one record, combining repeated fields into lists
            merged = _merge_records_to_single_row(records)
            columns = ["dataset_uuid"] + list(merged.keys())
            values = [dataset_uuid] + list(merged.values())
            placeholders = ", ".join(["%s"] * len(values))
            col_str = ", ".join(columns)
            cur.execute(
                f"INSERT INTO {table_name} ({col_str}) VALUES ({placeholders})",
                values,
            )
            return

        for record in records:
            if table_name == "subject":
                _insert_subject(cur, record, dataset_uuid, subject_id_to_uuid)
            elif table_name == "sample":
                _insert_sample(
                    cur, record, dataset_uuid, sample_mappings
                )
            else:
                # Generic dynamic insert (e.g. manifest)
                columns = ["dataset_uuid"] + list(record.keys())
                values = [dataset_uuid] + list(record.values())
                placeholders = ", ".join(["%s"] * len(values))
                col_str = ", ".join(columns)
                cur.execute(
                    f"INSERT INTO {table_name} ({col_str}) VALUES ({placeholders})",
                    values,
                )

    @staticmethod
    def _insert_dataset_mapping(
        cur,
        dataset_uuid: str,
        subject_id_to_uuid: dict,
        sample_mappings: list,
    ) -> None:
        """Populate ``dataset_mapping`` by pairing subjects with samples."""
        for sample_id, subject_id, sample_uuid in sample_mappings:
            subject_uuid = subject_id_to_uuid.get(subject_id) if subject_id else None
            if subject_uuid and sample_uuid:
                cur.execute(
                    "INSERT INTO dataset_mapping "
                    "(dataset_uuid, subject_uuid, sample_uuid, subject_id, sample_id) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (dataset_uuid, subject_uuid, sample_uuid, subject_id, sample_id),
                )
        logger.info("Inserted dataset_mapping rows for dataset %s", dataset_uuid)

    def upload_assay(self, assay_data):
        if self._postgres_enabled:
            self._postgres_uploader.upload_assay(assay_data)


# ------------------------------------------------------------------
# Module-level helper functions
# ------------------------------------------------------------------

def _merge_records_to_single_row(records: list[dict]) -> dict:
    """Merge multiple parsed Excel rows into a single dict.

    If a field appears in only one row its value is kept as-is (scalar).
    If a field has values across multiple rows the values are combined
    into a Python ``list``, which ``psycopg2`` maps to a Postgres array.
    """
    merged: dict = {}
    for record in records:
        for key, value in record.items():
            if key in merged:
                existing = merged[key]
                if isinstance(existing, list):
                    existing.append(value)
                else:
                    merged[key] = [existing, value]
            else:
                merged[key] = value
    return merged


def _insert_subject(
    cur, record: dict, dataset_uuid: str, subject_id_to_uuid: dict
) -> None:
    """Insert a single row into the ``subject`` table.

    The ``subject_uuid`` is auto-generated by the DB.  After insertion
    the mapping ``subject_id → subject_uuid`` is stored for later use
    in ``dataset_mapping``.
    """
    # Identify the subject_id from the record (used for mapping)
    subject_id = record.get("subject_id")

    # All keys in the record are now DB column names (from mapping)
    db_record = {k: v for k, v in record.items() if k != "subject_id"}
    columns = list(db_record.keys())
    values = list(db_record.values())
    placeholders = ", ".join(["%s"] * len(values))
    col_str = ", ".join(columns)

    if columns:
        cur.execute(
            f"INSERT INTO subject ({col_str}) VALUES ({placeholders}) RETURNING subject_uuid",
            values,
        )
    else:
        cur.execute("INSERT INTO subject DEFAULT VALUES RETURNING subject_uuid")

    subject_uuid = str(cur.fetchone()[0])
    if subject_id:
        subject_id_to_uuid[str(subject_id)] = subject_uuid


def _insert_sample(
    cur,
    record: dict,
    dataset_uuid: str,
    sample_mappings: list,
) -> None:
    """Insert a single row into the ``sample`` table.

    Captures mapping of ``sample_id`` -> ``subject_id`` -> ``sample_uuid``
    in the list for ``dataset_mapping`` population.
    """
    sample_id = record.get("sample_id")
    subject_id = record.get("subject_id") or record.get("was_derived_from_sample")

    db_record = {
        k: v
        for k, v in record.items()
        if k not in ("sample_id", "subject_id")
    }

    columns = list(db_record.keys())
    values = list(db_record.values())
    placeholders = ", ".join(["%s"] * len(values))
    col_str = ", ".join(columns)

    if columns:
        cur.execute(
            f"INSERT INTO sample ({col_str}) VALUES ({placeholders}) RETURNING sample_uuid",
            values,
        )
    else:
        cur.execute("INSERT INTO sample DEFAULT VALUES RETURNING sample_uuid")

    sample_uuid = str(cur.fetchone()[0])
    if sample_id and subject_id:
        sample_mappings.append((str(sample_id), str(subject_id), sample_uuid))
